import sqlite3
import openpyxl
from pathlib import Path


def create_database():
    conn = sqlite3.connect('obraz_plus.db')
    cursor = conn.cursor()

    # Таблица типов материалов
    cursor.execute('''
                   CREATE TABLE IF NOT EXISTS MaterialTypes
                   (
                       id
                       INTEGER
                       PRIMARY
                       KEY
                       AUTOINCREMENT,
                       name
                       TEXT
                       NOT
                       NULL,
                       loss_percentage
                       REAL
                       NOT
                       NULL
                   )
                   ''')

    # Таблица материалов
    cursor.execute('''
                   CREATE TABLE IF NOT EXISTS Materials
                   (
                       id
                       INTEGER
                       PRIMARY
                       KEY
                       AUTOINCREMENT,
                       name
                       TEXT
                       NOT
                       NULL,
                       type_id
                       INTEGER,
                       unit_price
                       REAL
                       NOT
                       NULL
                       CHECK
                   (
                       unit_price
                       >=
                       0
                   ),
                       stock_quantity REAL NOT NULL,
                       min_quantity REAL NOT NULL CHECK
                   (
                       min_quantity
                       >=
                       0
                   ),
                       package_quantity REAL NOT NULL,
                       unit_of_measure TEXT NOT NULL,
                       FOREIGN KEY
                   (
                       type_id
                   ) REFERENCES MaterialTypes
                   (
                       id
                   )
                       )
                   ''')

    # Таблица типов продукции
    cursor.execute('''
                   CREATE TABLE IF NOT EXISTS ProductTypes
                   (
                       id
                       INTEGER
                       PRIMARY
                       KEY
                       AUTOINCREMENT,
                       name
                       TEXT
                       NOT
                       NULL,
                       coefficient
                       REAL
                       NOT
                       NULL
                   )
                   ''')

    # Таблица продукции
    cursor.execute('''
                   CREATE TABLE IF NOT EXISTS Products
                   (
                       id
                       INTEGER
                       PRIMARY
                       KEY
                       AUTOINCREMENT,
                       name
                       TEXT
                       NOT
                       NULL,
                       article
                       TEXT
                       NOT
                       NULL
                       UNIQUE,
                       min_partner_price
                       REAL
                       NOT
                       NULL,
                       type_id
                       INTEGER,
                       FOREIGN
                       KEY
                   (
                       type_id
                   ) REFERENCES ProductTypes
                   (
                       id
                   )
                       )
                   ''')

    # Таблица связи материалов и продукции
    cursor.execute('''
                   CREATE TABLE IF NOT EXISTS ProductMaterials
                   (
                       product_id
                       INTEGER,
                       material_id
                       INTEGER,
                       required_quantity
                       REAL
                       NOT
                       NULL,
                       PRIMARY
                       KEY
                   (
                       product_id,
                       material_id
                   ),
                       FOREIGN KEY
                   (
                       product_id
                   ) REFERENCES Products
                   (
                       id
                   ),
                       FOREIGN KEY
                   (
                       material_id
                   ) REFERENCES Materials
                   (
                       id
                   )
                       )
                   ''')

    conn.commit()
    conn.close()


def import_from_excel():
    conn = sqlite3.connect('obraz_plus.db')
    cursor = conn.cursor()

    # Очистка таблиц в правильном порядке
    cursor.execute("DELETE FROM ProductMaterials")
    cursor.execute("DELETE FROM Products")
    cursor.execute("DELETE FROM Materials")
    cursor.execute("DELETE FROM ProductTypes")
    cursor.execute("DELETE FROM MaterialTypes")

    # Импорт типов материалов
    try:
        wb = openpyxl.load_workbook('Material_type_import.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cursor.execute('INSERT INTO MaterialTypes (name, loss_percentage) VALUES (?, ?)', row)
    except Exception as e:
        print(f"Ошибка при импорте типов материалов: {e}")

    # Импорт материалов
    try:
        wb = openpyxl.load_workbook('Materials_import.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            material_name, type_name, *rest = row
            cursor.execute('SELECT id FROM MaterialTypes WHERE name = ?', (type_name,))
            type_id = cursor.fetchone()
            if type_id:
                cursor.execute('''
                               INSERT INTO Materials (name, type_id, unit_price, stock_quantity,
                                                      min_quantity, package_quantity, unit_of_measure)
                               VALUES (?, ?, ?, ?, ?, ?, ?)
                               ''', (material_name, type_id[0], *rest))
            else:
                print(f"Тип материала '{type_name}' не найден для материала '{material_name}'")
    except Exception as e:
        print(f"Ошибка при импорте материалов: {e}")

    # Импорт типов продукции
    try:
        wb = openpyxl.load_workbook('Product_type_import.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cursor.execute('INSERT INTO ProductTypes (name, coefficient) VALUES (?, ?)', row)
    except Exception as e:
        print(f"Ошибка при импорте типов продукции: {e}")

    # Импорт продукции
    try:
        wb = openpyxl.load_workbook('Products_import.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            type_name, product_name, article, price = row
            cursor.execute('SELECT id FROM ProductTypes WHERE name = ?', (type_name,))
            type_id = cursor.fetchone()
            if type_id:
                cursor.execute('''
                               INSERT INTO Products (name, article, min_partner_price, type_id)
                               VALUES (?, ?, ?, ?)
                               ''', (product_name, article, price, type_id[0]))
            else:
                print(f"Тип продукции '{type_name}' не найден для продукта '{product_name}'")
    except Exception as e:
        print(f"Ошибка при импорте продукции: {e}")

    # Импорт связей материалов и продукции
    try:
        wb = openpyxl.load_workbook('Material_products__import.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            material_name, product_name, quantity = row
            cursor.execute('SELECT id FROM Materials WHERE name = ?', (material_name,))
            material_id = cursor.fetchone()
            cursor.execute('SELECT id FROM Products WHERE name = ?', (product_name,))
            product_id = cursor.fetchone()

            if material_id and product_id:
                cursor.execute('''
                               INSERT INTO ProductMaterials (product_id, material_id, required_quantity)
                               VALUES (?, ?, ?)
                               ''', (product_id[0], material_id[0], quantity))
            else:
                if not material_id:
                    print(f"Материал '{material_name}' не найден")
                if not product_id:
                    print(f"Продукт '{product_name}' не найден")
    except Exception as e:
        print(f"Ошибка при импорте связей материалов и продукции: {e}")

    conn.commit()
    conn.close()


def calculate_product_quantity(product_type_id, material_type_id, raw_quantity, param1, param2):
    """Расчет количества продукции из сырья с учетом потерь"""
    conn = sqlite3.connect('obraz_plus.db')
    cursor = conn.cursor()

    try:
        # Получаем коэффициент продукта и процент потерь
        cursor.execute('SELECT coefficient FROM ProductTypes WHERE id = ?', (product_type_id,))
        product_coeff = cursor.fetchone()[0]

        cursor.execute('SELECT loss_percentage FROM MaterialTypes WHERE id = ?', (material_type_id,))
        loss_percent = cursor.fetchone()[0]

        # Расчет сырья на 1 единицу продукции
        raw_per_unit = param1 * param2 * product_coeff

        # Учет потерь
        effective_raw = raw_quantity * (1 - loss_percent / 100)

        # Расчет количества продукции
        product_quantity = effective_raw / raw_per_unit

        return int(product_quantity) if product_quantity >= 0 else -1
    except Exception as e:
        print(f"Ошибка расчета: {e}")
        return -1
    finally:
        conn.close()


if __name__ == "__main__":
    create_database()
    import_from_excel()